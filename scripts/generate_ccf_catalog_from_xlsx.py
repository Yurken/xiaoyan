#!/usr/bin/env python3

import argparse
import json
import re
from pathlib import Path

from openpyxl import load_workbook


def clean(value: object | None) -> str:
    if value is None:
        return ""
    text = str(value).replace("\u3000", " ").replace("\n", " ")
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def looks_like_url(value: str) -> bool:
    return value.startswith("http://") or value.startswith("https://")


def normalize_aliases(label: str, full_name: str, url: str) -> list[str]:
    aliases: list[str] = []
    seen: set[str] = set()

    def add(value: str) -> None:
        normalized = clean(value)
        if not normalized:
            return
        key = normalized.lower()
        if key in seen:
            return
        seen.add(key)
        aliases.append(normalized)

    add(label)
    add(full_name)
    add(re.sub(r"\s*\([^)]*\)", "", full_name))
    add(full_name.replace("Proceedings of the ", ""))
    add(full_name.replace("Proceedings of ", ""))
    add(full_name.replace("International Conference on ", ""))
    add(full_name.replace("IEEE International Conference on ", ""))
    add(full_name.replace("ACM International Conference on ", ""))

    # Derive short alias from SIG* labels, e.g. SIGKDD -> KDD
    if label.upper().startswith("SIG") and len(label) > 3:
        short = label[3:]
        if short and short[0].isalpha():
            add(short)

    # Extract DBLP key from URL as alias, e.g. …/db/conf/kdd/ -> KDD
    m = re.search(r"db/(?:conf|journals)/([^/]+)", url, re.IGNORECASE)
    if m:
        dblp_key = m.group(1).upper()
        add(dblp_key)

    return aliases


def extract_rating(value: str) -> str:
    match = re.search(r"([ABC])类", value)
    return match.group(1) if match else ""


def parse_entries(xlsx_path: Path) -> list[dict[str, object]]:
    workbook = load_workbook(xlsx_path, data_only=True)
    sheet = workbook.active

    current_kind = ""
    current_area = ""
    current_rating = ""
    entries: list[dict[str, object]] = []

    for row in sheet.iter_rows(values_only=True):
        cells = [clean(cell) for cell in row]
        nonempty = [cell for cell in cells if cell]
        if not nonempty:
            continue

        first = nonempty[0]

        if "中国计算机学会推荐国际学术期刊" in nonempty:
            current_kind = "journal"
            current_area = ""
            current_rating = ""
            continue

        if "中国计算机学会推荐国际学术会议" in nonempty:
            current_kind = "conference"
            current_area = ""
            current_rating = ""
            continue

        if (
            current_kind
            and len(first) >= 4
            and first[0] in {"(", "（"}
            and first[-1] in {")", "）"}
            and not re.search(r"\d{4}", first)
        ):
            current_area = first.strip("()（）")
            continue

        if "类" in first and any(letter in first for letter in ("A", "B", "C")):
            current_rating = extract_rating(first)
            continue

        if first == "序号":
            continue

        serial = clean(cells[0])
        if not serial.isdigit() or not current_kind or not current_area or not current_rating:
            continue

        url = next((value for value in reversed(nonempty) if looks_like_url(value)), "")
        if not url:
            continue

        data_cells = [value for value in nonempty[1:] if value != url]
        if len(data_cells) < 2:
            continue

        publisher = data_cells[-1]
        name_cells = data_cells[:-1]
        if not name_cells:
            continue

        if len(name_cells) == 1:
            label = ""
            full_name = name_cells[0]
        else:
            label = name_cells[0]
            full_name = " ".join(name_cells[1:])

        entry = {
            "kind": current_kind,
            "rating": current_rating,
            "area": current_area,
            "label": clean(label),
            "full_name": clean(full_name),
            "publisher": clean(publisher),
            "url": url,
            "aliases": normalize_aliases(label, full_name, url),
        }

        if entry["full_name"]:
            entries.append(entry)

    return entries


def main() -> None:
    parser = argparse.ArgumentParser(description="Generate ResearchCopilot CCF catalog JSON from the official XLSX.")
    parser.add_argument("xlsx", type=Path)
    parser.add_argument("output", type=Path)
    args = parser.parse_args()

    entries = parse_entries(args.xlsx)
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(json.dumps(entries, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(f"generated {len(entries)} entries -> {args.output}")


if __name__ == "__main__":
    main()
